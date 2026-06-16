import json
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from accounts.decorators import role_required
from accounts.models import User
from .forms import BatchForm, FileUploadForm, HolidayForm, TaskForm
from .models import Batch, FileUpload, Holiday, Task
from .services import (
    TASK_TYPE_COLORS,
    build_calendar_events,
    get_scheduler_events,
    get_task_type_color,
    get_visible_batches,
)


@login_required
def task_list(request):
    queryset = Task.objects.select_related("trainer")
    if request.user.role == "trainer":
        queryset = queryset.filter(trainer=request.user)
    tasks = queryset.order_by("-date")
    return render(request, "tasks/task_list.html", {"tasks": tasks})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def create_task(request):
    if request.method == "POST":
        form = TaskForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Task entry saved.")
            return redirect("tasks:task_list")
    else:
        form = TaskForm()
    return render(request, "tasks/task_form.html", {"form": form})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def delete_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if request.method == "POST":
        task.delete()
        messages.success(request, "Task deleted.")
        return redirect("tasks:task_list")
    return render(request, "tasks/task_confirm_delete.html", {"task": task})


@login_required
@role_required(User.Role.TRAINER)
def upload_file(request):
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.trainer = request.user
            upload.save()
            messages.success(request, "File uploaded successfully.")
            return redirect("tasks:file_list")
    else:
        form = FileUploadForm()
    return render(request, "tasks/file_form.html", {"form": form})


@login_required
def file_list(request):
    uploads = FileUpload.objects.filter(trainer=request.user) if request.user.role == "trainer" else FileUpload.objects.all()
    return render(
        request,
        "tasks/file_list.html",
        {"uploads": uploads, "form": FileUploadForm(), "can_upload_files": request.user.role == "trainer"},
    )


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def batch_list(request):
    batches = Batch.objects.select_related("trainer", "circle").prefetch_related("holidays")
    return render(request, "tasks/batch_list.html", {"batches": batches})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def create_batch(request):
    if request.method == "POST":
        form = BatchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Batch created.")
            return redirect("tasks:batch_list")
    else:
        form = BatchForm()
    return render(request, "tasks/batch_form.html", {"form": form})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def delete_batch(request, batch_id):
    batch = get_object_or_404(Batch, pk=batch_id)
    if request.method == "POST":
        batch.delete()
        messages.success(request, "Batch deleted.")
        return redirect("tasks:batch_list")
    return render(request, "tasks/batch_confirm_delete.html", {"batch": batch})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def holiday_list(request):
    holidays = Holiday.objects.all()
    return render(request, "tasks/holiday_list.html", {"holidays": holidays})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def create_holiday(request):
    if request.method == "POST":
        form = HolidayForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Holiday added.")
            return redirect("tasks:holiday_list")
    else:
        form = HolidayForm()
    return render(request, "tasks/holiday_form.html", {"form": form})


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def delete_holiday(request, holiday_id):
    holiday = get_object_or_404(Holiday, pk=holiday_id)
    if request.method == "POST":
        holiday.delete()
        messages.success(request, "Holiday deleted.")
        return redirect("tasks:holiday_list")
    return render(request, "tasks/holiday_confirm_delete.html", {"holiday": holiday})


@login_required
def calendar_feed(request):
    visible_batches = get_visible_batches(request.user)
    visible_trainers = User.objects.filter(role=User.Role.TRAINER)
    if request.user.role == User.Role.TRAINER:
        visible_trainers = visible_trainers.filter(pk=request.user.pk)
    elif request.user.role == User.Role.CIRCLE_LEAD:
        visible_trainers = visible_trainers.filter(trainer_profile__circle__manager=request.user).distinct()
    return render(
        request,
        "tasks/calendar.html",
        {
            "calendar_events_json": json.dumps(build_calendar_events(request.user)),
            "manager_schedule_enabled": request.user.role == "manager",
            "batch_filters": visible_batches,
            "trainer_filters": visible_trainers,
            "task_legend": [
                {"name": "Training", "color": get_task_type_color("training")},
                {"name": "Assessment", "color": get_task_type_color("assessment")},
                {"name": "Free Slot", "color": get_task_type_color("free")},
                {"name": "Deck Preparation", "color": get_task_type_color("deck")},
                {"name": "Project Support", "color": get_task_type_color("project")},
                {"name": "Leave", "color": get_task_type_color("leave")},
                {"name": "Holiday", "color": get_task_type_color("holiday")},
                {"name": "Meeting", "color": get_task_type_color("meeting")},
            ],
            "can_export_timetable": request.user.role in {User.Role.ADMIN, User.Role.MANAGER},
            "today_str": timezone.localdate().isoformat(),
        },
    )


@login_required
def get_schedule_by_date(request, date):
    if request.user.role != "manager":
        return HttpResponseForbidden("Managers only.")

    target_date = parse_date(date)
    if not target_date:
        return JsonResponse({"detail": "Invalid date format."}, status=400)

    schedules = (
        Task.objects.filter(date=target_date)
        .select_related("trainer")
        .order_by("trainer__name", "task_type")
    )
    payload = [
        {
            "trainer": task.trainer.name,
            "task": task.get_task_type_display(),
            "hours": float(task.hours),
            "status": task.get_status_display(),
        }
        for task in schedules
    ]
    return JsonResponse(payload, safe=False)


@login_required
def scheduler_events_feed(request):
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    batch_id = request.GET.get("batch")
    trainer_id = request.GET.get("trainer")

    start_date = datetime.fromisoformat(start_str).date() if start_str else None
    end_date = datetime.fromisoformat(end_str).date() if end_str else None

    events = get_scheduler_events(
        request.user,
        start_date=start_date,
        end_date=end_date,
        batch_id=batch_id or None,
        trainer_id=trainer_id or None,
    )
    return JsonResponse(events, safe=False)


@login_required
def scheduler_day_timetable(request, date):
    target_date = parse_date(date)
    if not target_date:
        return JsonResponse({"detail": "Invalid date format."}, status=400)

    batch_id = request.GET.get("batch")
    trainer_id = request.GET.get("trainer")
    events = get_scheduler_events(
        request.user,
        start_date=target_date,
        end_date=target_date,
        batch_id=batch_id or None,
        trainer_id=trainer_id or None,
    )
    rows = [
        {
            "trainer": event["extendedProps"]["trainer"],
            "batch": event["extendedProps"]["batch"],
            "task_type": event["extendedProps"]["task_type"],
            "start_time": event["extendedProps"]["start_time"],
            "end_time": event["extendedProps"]["end_time"],
            "occupancy_status": event["extendedProps"]["occupancy_status"],
            "duration": event["extendedProps"]["duration"],
            "circle": event["extendedProps"]["circle"],
            "description": event["extendedProps"]["description"],
            "batch_color": event["extendedProps"]["batch_color"],
            "task_color": event["extendedProps"]["task_color"],
        }
        for event in events
    ]
    return JsonResponse(rows, safe=False)


@login_required
@role_required(User.Role.ADMIN, User.Role.MANAGER)
def export_daily_timetable(request):
    date_str = request.GET.get("date")
    batch_id = request.GET.get("batch")
    trainer_id = request.GET.get("trainer")
    target_date = parse_date(date_str) if date_str else timezone.localdate()
    if not target_date:
        return JsonResponse({"detail": "Invalid date."}, status=400)

    events = get_scheduler_events(
        request.user,
        start_date=target_date,
        end_date=target_date,
        batch_id=batch_id or None,
        trainer_id=trainer_id or None,
    )
    rows = [event["extendedProps"] for event in events if event.get("display") != "background"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily Timetable"
    headers = [
        "Trainer",
        "Batch",
        "Task Type",
        "Start Time",
        "End Time",
        "Occupancy Status",
        "Occupancy %",
        "Duration (hrs)",
        "Circle",
        "Description",
    ]
    sheet.append(headers)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for index, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row in enumerate(rows, start=2):
        sheet.append(
            [
                row.get("trainer", "-"),
                row.get("batch", "-"),
                row.get("task_type", "-"),
                row.get("start_time", "-"),
                row.get("end_time", "-"),
                row.get("occupancy_status", "-"),
                row.get("occupancy_percent", 0),
                row.get("duration", 0),
                row.get("circle", "-"),
                row.get("description", "-"),
            ]
        )
        task_color = str(row.get("task_color", "#2563eb")).replace("#", "").upper()
        if len(task_color) == 6:
            fill = PatternFill(start_color=task_color, end_color=task_color, fill_type="solid")
            sheet.cell(row=row_index, column=3).fill = fill

    column_widths = [22, 20, 18, 14, 14, 18, 12, 14, 20, 34]
    for col, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + col)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="daily_timetable_{target_date.isoformat()}.xlsx"'
    workbook.save(response)
    return response
